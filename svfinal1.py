import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the input file and the reference stock file
input_file = 'gen.xlsx'
stock_file = 'stockgen.xlsx'
output_file = 'svgen.xlsx'

# Read the Excel files, specifying 'accno' as string to preserve leading zeros
df_input = pd.read_excel(input_file, dtype={'accno': str})
df_stock = pd.read_excel(stock_file, dtype={'accno': str})

# Create a column 'Status' to mark the status of each 'accno'
df_input['Status'] = ''

# Check for duplicates in the input file
df_input['Is_Duplicate'] = df_input.duplicated(subset=['accno'], keep=False)

# Create lists to hold missing and duplicate accnos
missing_accnos = []
duplicate_accnos = []

# Check each accno in the input file against the stock file
for idx, row in df_input.iterrows():
    accno = row['accno']
    if accno in df_stock['accno'].values:
        if df_input[df_input['accno'] == accno].shape[0] > 1:
            df_input.at[idx, 'Status'] = 'Duplicate'
            duplicate_accnos.append(accno)
        else:
            df_input.at[idx, 'Status'] = 'Verified'
    else:
        df_input.at[idx, 'Status'] = 'Missing'
        missing_accnos.append(accno)

# Find stock accnos not present in the input file
missing_in_input = df_stock[~df_stock['accno'].isin(df_input['accno'])]['accno'].tolist()

# Save the main DataFrame to a new Excel file
df_input.to_excel(output_file, index=False)

# Create DataFrames for missing and duplicate accnos
df_missing = pd.DataFrame(missing_accnos, columns=['missing accnos'])
df_duplicates = pd.DataFrame(duplicate_accnos, columns=['duplicate accnos'])
df_missing_in_input = pd.DataFrame(missing_in_input, columns=['missing in stock'])

# Load the workbook and worksheet to apply cell formatting
wb = load_workbook(output_file)
ws = wb.active

# Define the colors for the cells
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Apply the color formatting based on the 'Status' column
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    status_cell = row[-1]  # The last column is 'Status'
    if status_cell.value == 'Verified':
        for cell in row:
            cell.fill = green_fill
    elif status_cell.value == 'Duplicate':
        for cell in row:
            cell.fill = yellow_fill
    elif status_cell.value == 'Missing':
        for cell in row:
            cell.fill = red_fill

# Save the changes to the main sheet
wb.save(output_file)

# Reopen the workbook in append mode to add new sheets
with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
    df_missing.to_excel(writer, sheet_name='missing accnos', index=False)
    df_duplicates.to_excel(writer, sheet_name='duplicate accnos', index=False)
    df_missing_in_input.to_excel(writer, sheet_name='outside the series', index=False)

print(f"Processed file saved as {output_file}")
